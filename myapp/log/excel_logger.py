import os
from datetime import datetime
from threading import Lock
from openpyxl import Workbook, load_workbook
from django.conf import settings

CONTEXT_COLUMNS = [
    "search_id",
    "timestamp",
    "halte_asal",
    "halte_tujuan",
    "tanggal_berangkat",
    "jam_berangkat",
    "preferensi",
    "metode_solver",
    "mode_waktu",
    "param_speed",
    "weights_waktu",
    "weights_biaya",
    "weights_transit",
]

RESULT_COLUMNS = ["waktu", "jarak", "biaya", "transit", "obj_value", "runtime"]

HEADERS = CONTEXT_COLUMNS + RESULT_COLUMNS  # 19 columns

_LOG_DIR = os.path.join(settings.BASE_DIR, "myapp", "static", "data", "log")
LOG_PATH = os.path.join(_LOG_DIR, "search_log.xlsx")

_write_lock = Lock()


def _ensure_workbook():
    os.makedirs(_LOG_DIR, exist_ok=True)
    if not os.path.exists(LOG_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "log"
        ws.append(HEADERS)
        wb.save(LOG_PATH)


def _next_search_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if isinstance(val, int) and val > max_id:
            max_id = val
    return max_id + 1


def append_search(meta, result):
    """Append one row representing a single-solver search.

    meta: dict with keys halte_asal, halte_tujuan, tanggal_berangkat (date),
          jam_berangkat (time), preferensi, metode_solver, mode_waktu,
          param_speed, weights (dict with waktu/biaya/transit).
    result: dict with keys waktu_tempuh_menit, jarak_km, total_biaya,
            jumlah_transit, z_score, runtime_sec.

    Returns the assigned search_id.
    """
    _ensure_workbook()
    with _write_lock:
        wb = load_workbook(LOG_PATH)
        ws = wb.active

        search_id = _next_search_id(ws)
        weights = meta.get("weights", {}) or {}

        row = [
            search_id,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            meta.get("halte_asal", ""),
            meta.get("halte_tujuan", ""),
            str(meta.get("tanggal_berangkat", "")),
            str(meta.get("jam_berangkat", "")),
            meta.get("preferensi", ""),
            meta.get("metode_solver", ""),
            meta.get("mode_waktu", ""),
            meta.get("param_speed", ""),
            weights.get("waktu", ""),
            weights.get("biaya", ""),
            weights.get("transit", ""),
            _round(result.get("waktu_tempuh_menit"), 4),
            _round(result.get("jarak_km"), 4),
            _round(result.get("total_biaya"), 2),
            _to_int(result.get("jumlah_transit")),
            _round(result.get("z_score"), 4),
            _round(result.get("runtime_sec"), 4),
        ]

        ws.append(row)
        wb.save(LOG_PATH)
        return search_id


def read_log(newest_first=True):
    """Return list of dicts (one per row), header keys preserved."""
    if not os.path.exists(LOG_PATH):
        return []
    wb = load_workbook(LOG_PATH, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    out = []
    for raw in rows_iter:
        if raw is None or all(v is None or v == "" for v in raw):
            continue
        out.append(dict(zip(HEADERS, raw)))
    wb.close()
    if newest_first:
        out.reverse()
    return out


def _round(v, ndigits):
    if v is None or v == "":
        return ""
    try:
        return round(float(v), ndigits)
    except (TypeError, ValueError):
        return ""


def _to_int(v):
    if v is None or v == "":
        return ""
    try:
        return int(v)
    except (TypeError, ValueError):
        return ""
